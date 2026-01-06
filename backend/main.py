from fastapi import FastAPI, Depends, HTTPException, Header
from sqlalchemy.orm import Session
from fastapi.middleware.cors import CORSMiddleware
from datetime import date, timedelta
import models,schemas,string
from database import engine, get_db, SessionLocal
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import pandas as pd
import random
from typing import List

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="智慧酒店管理系统后端")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有来源，开发环境比较方便
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def verify_token(token: str = Header(None)):
    if token != "fake-jwt-token":
        raise HTTPException(status_code=401, detail="未登录或登录已过期")
    return token


# --- 辅助工具函数 ---
def generate_random_name():
    first_names = ["张", "王", "李", "刘", "赵", "陈", "杨", "周", "吴", "徐", "孙", "马", "朱", "胡", "郭", "何", "高",
                   "林"]
    last_names = ["伟", "芳", "娜", "秀兰", "洋", "敏", "静", "杰", "强", "涛", "丽", "艳", "帅", "磊", "军", "勇",
                  "丹"]
    name = random.choice(first_names) + random.choice(last_names)
    if random.random() > 0.5:
        name += random.choice(last_names)
    return name


def generate_random_phone():
    prefixes = ["138", "139", "150", "188", "177", "131", "135"]
    return random.choice(prefixes) + "".join(random.choices(string.digits, k=8))


def generate_random_id_card():
    area = random.randint(110000, 650000)
    birthday = f"{random.randint(1980, 2005)}{random.randint(1, 12):02d}{random.randint(1, 28):02d}"
    suffix = "".join(random.choices(string.digits, k=4))
    return f"{area}{birthday}{suffix}"


# --- 主初始化函数 ---
def init_db_data(db: Session):
    if db.query(models.User).filter(models.User.username == "admin").first():
        return

    print("🚀 正在按照最新模型初始化演示数据...")

    # 1. 创建管理员
    db.add(models.User(username="admin", password="123"))

    # 2. 房型模板
    templates = [
        {"type": "影音大床房", "config": "120寸投影, 5.1音响, 芝华仕沙发", "price": 388},
        {"type": "电竞双人间", "config": "RTX4090显卡, 240Hz显示器, 电竞椅", "price": 488},
        {"type": "商务麻将房", "config": "自动麻将机, 功夫茶具, 隔音处理", "price": 588},
        {"type": "标准双床房", "config": "两张1.5米床, 独立卫浴, 办公桌", "price": 199}
    ]

    # 3. 创建房间
    rooms = []
    for floor in [1, 2, 3]:
        for i in range(1, 6):
            tpl = random.choice(templates)
            room = models.Room(
                number=f"{floor}0{i}",
                room_type=tpl["type"],
                configuration=tpl["config"],
                price=tpl["price"],
                status="空闲"
            )
            db.add(room)
            rooms.append(room)
    db.flush()

    # 4. 初始化会员 (适配你的 Member 模型)
    members = []
    levels = ["普通会员", "白金会员", "钻石会员"]
    for _ in range(20):
        m_name = generate_random_name()
        m_phone = generate_random_phone()
        member = models.Member(
            name=m_name,
            phone=m_phone,
            password="123",  # 你的模型有这个字段，初始化默认给123
            level=random.choice(levels),
            points=random.randint(100, 5000),
            balance=float(random.randint(0, 2000)),
            reg_date=date.today() - timedelta(days=random.randint(1, 365))  # 适配你的字段名 reg_date
        )
        db.add(member)
        members.append(member)
    db.flush()

    # 5. 初始化预订与住客同步
    today = date.today()
    for _ in range(50):
        room = random.choice(rooms)
        start_date = today + timedelta(days=random.randint(-15, 10))
        end_date = start_date + timedelta(days=random.randint(1, 4))

        # 模拟会员/散客分配身份证（Room表需要身份证字段，Member表不需要）
        is_member = random.random() > 0.4
        current_id_card = generate_random_id_card()  # 无论是不是会员，入住都要身份证

        if is_member:
            m = random.choice(members)
            g_name, g_phone = m.name, m.phone
        else:
            g_name, g_phone = generate_random_name(), generate_random_phone()

        status = "待入住"
        if end_date < today:
            status = "已离店/完成"
        elif start_date <= today <= end_date:
            status = "入住中"
            # 同步更新 Room 表 (Room模型有这些字段)
            room.status = "已入住"
            room.guest_name = g_name
            room.guest_phone = g_phone
            room.guest_id_card = current_id_card

        db.add(models.Booking(
            room_id=room.id,
            guest_name=g_name,
            start_date=start_date,
            end_date=end_date,
            status=status
        ))

    db.commit()
    print("✅ 数据初始化成功！")


@app.on_event("startup")
def startup_event():
    db = SessionLocal()
    try:
        init_db_data(db)
    finally:
        db.close()


@app.post("/api/login")
def login(request: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == request.username).first()
    if not user or user.password != request.password:
        raise HTTPException(status_code=400, detail="用户名或密码错误")
    return {"message": "登录成功", "token": "fake-jwt-token", "username": user.username}


# @app.post("/api/init")
# def init_data(db: Session = Depends(get_db)):
#     if db.query(models.User).count() == 0:
#         db.add(models.User(username="admin", password="123"))
#     if db.query(models.Room).count() == 0:
#         db.add(models.Room(number="101", room_type="标准间", price=199.0))
#         db.add(models.Room(number="201", room_type="豪华大床房", price=399.0))
#     db.commit()
#     return {"msg": "初始化成功"}


@app.get("/api/rooms")
def read_rooms(db: Session = Depends(get_db)):
    return db.query(models.Room).all()


@app.post("/api/rooms")
def create_room(room: schemas.RoomCreate, db: Session = Depends(get_db)):
    db_room = db.query(models.Room).filter(models.Room.number == room.number).first()
    if db_room:
        raise HTTPException(status_code=400, detail="房间号已存在")
    new_room = models.Room(**room.dict())
    db.add(new_room)
    db.commit()
    db.refresh(new_room)
    return new_room


@app.delete("/api/rooms/{room_id}")
def delete_room(room_id: int, db: Session = Depends(get_db)):
    db_room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not db_room:
        raise HTTPException(status_code=404, detail="房间未找到")
    db.delete(db_room)
    db.commit()
    return {"message": "房间删除成功"}


@app.put("/api/rooms/{room_id}/status")
def update_room_status(room_id: int, data: schemas.RoomStatusUpdate, db: Session = Depends(get_db)):
    room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")

    room.status = data.status
    room.guest_name = data.guest_name
    room.guest_id_card = data.guest_id_card
    room.guest_phone = data.guest_phone

    db.commit()
    return {"message": "更新成功"}

@app.get("/api/rooms/available")
def get_available_rooms(target_date: date, db: Session = Depends(get_db)):

    occupied_room_ids = db.query(models.Booking.room_id).filter(
        models.Booking.start_date <= target_date,
        models.Booking.end_date >= target_date,
        models.Booking.status != "已取消"
    ).all()

    occupied_ids = [r[0] for r in occupied_room_ids]

    available_rooms = db.query(models.Room).filter(~models.Room.id.in_(occupied_ids)).all()
    return available_rooms


@app.post("/api/bookings")
def create_booking(booking: schemas.BookingCreate, db: Session = Depends(get_db)):
    conflict = db.query(models.Booking).filter(
        models.Booking.room_id == booking.room_id,
        models.Booking.status != "已取消",
        models.Booking.start_date < booking.end_date,
        models.Booking.end_date > booking.start_date
    ).first()

    if conflict:
        raise HTTPException(status_code=400, detail="该时间段房间已被占用")

    new_booking = models.Booking(
        room_id=booking.room_id,
        guest_name=booking.guest_name,
        start_date=booking.start_date,
        end_date=booking.end_date,
        status="待入住"
    )
    db.add(new_booking)
    db.commit()
    db.refresh(new_booking)
    return {"message": "预订成功", "id": new_booking.id}

@app.get("/api/bookings/today")
def get_today_bookings(db: Session = Depends(get_db)):
    from datetime import date
    today = date.today()


    bookings = db.query(models.Booking).filter(
        models.Booking.start_date == today,
        models.Booking.status == "待入住"
    ).all()

    result = []
    for b in bookings:
        room = db.query(models.Room).filter(models.Room.id == b.room_id).first()
        result.append({
            "guest_name": b.guest_name,
            "room_number": room.number if room else "未知",
            "time": "14:00 入住"
        })
    return result


@app.put("/api/rooms/{room_id}")
def update_room_info(room_id: int, room_data: schemas.RoomUpdate, db: Session = Depends(get_db)):
    db_room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not db_room:
        raise HTTPException(status_code=404, detail="未找到该房间")

    # 1. 更新基本字段
    db_room.room_type = room_data.room_type
    db_room.price = room_data.price

    # 2. 更新配置字段 (现在 room_data.configuration 一定有值或为 None)
    db_room.configuration = room_data.configuration

    db.commit()
    db.refresh(db_room)  # 刷新对象，确保返回的是数据库最新状态
    return {"message": "更新成功", "data": db_room}


from sqlalchemy import func


@app.get("/api/reports/stats")
def get_report_stats(db: Session = Depends(get_db)):
    today = date.today()

    # 1. 营收统计：基于房间单价和预订天数统计
    today_revenue = db.query(func.sum(models.Room.price)).join(
        models.Booking, models.Room.id == models.Booking.room_id
    ).filter(
        models.Booking.start_date <= today,
        models.Booking.end_date >= today,
        models.Booking.status != "已取消"
    ).scalar() or 0

    # 2. 入住率统计
    total_rooms = db.query(models.Room).count()
    occupied_rooms = db.query(models.Room).filter(models.Room.status == "已入住").count()
    occupancy_rate = round((occupied_rooms / total_rooms * 100), 1) if total_rooms > 0 else 0

    # 3. 待处理预订
    pending_bookings = db.query(models.Booking).filter(
        models.Booking.status == "待入住",
        models.Booking.start_date >= today
    ).count()

    return [
        {"title": "今日预计营收", "value": f"{today_revenue:.2f}", "prefix": "¥"},
        {"title": "当前入住率", "value": str(occupancy_rate), "prefix": "%"},
        {"title": "待处理预订", "value": str(pending_bookings), "prefix": ""},
        {"title": "客房总数", "value": str(total_rooms), "prefix": ""}
    ]


# --- 新增/优化：真实数据统计图表接口 ---
@app.get("/api/reports/chart")
def get_chart_data(db: Session = Depends(get_db)):
    days_labels = []
    counts_data = []

    # 循环过去 7 天
    for i in range(6, -1, -1):
        target_date = date.today() - timedelta(days=i)
        # 格式化日期作为坐标轴标签 (例如: 01-05)
        days_labels.append(target_date.strftime("%m-%d"))

        # 统计当天的订单数量
        count = db.query(models.Booking).filter(
            models.Booking.start_date == target_date,
            models.Booking.status != "已取消"
        ).count()
        counts_data.append(count)

    return {
        "days": days_labels,
        "data": counts_data
    }


@app.get("/api/reports/room-type-dist")
def get_room_type_distribution(db: Session = Depends(get_db)):
    # 统计不同房型的订单分布
    results = db.query(
        models.Room.room_type,
        func.count(models.Booking.id).label('count')
    ).join(models.Booking, models.Room.id == models.Booking.room_id).group_by(models.Room.room_type).all()

    return [{"name": r.room_type, "value": r.count} for r in results]


# --- 新增：实时房态墙数据 ---
# --- 修改后的实时房态墙数据接口 ---
@app.get("/api/reports/room-wall")
def get_room_wall(db: Session = Depends(get_db)):
    rooms = db.query(models.Room).all()
    return [{
        "id": r.id,
        "number": r.number,
        "room_type": r.room_type, # 统一使用 room_type
        "status": r.status,
        "price": r.price,
        "configuration": r.configuration  # 必须返回配置，前端才能显示
    } for r in rooms]


from urllib.parse import quote
from datetime import datetime


@app.get("/api/reports/export-excel")
def export_excel(db: Session = Depends(get_db)):
    # 1. 获取当前系统日期
    today = date.today()

    # 2. 从数据库抓取原始数据
    bookings = db.query(
        models.Booking.id,
        models.Booking.guest_name,
        models.Room.number.label("room_number"),
        models.Booking.start_date,
        models.Booking.end_date,
        models.Booking.status
    ).join(models.Room, models.Room.id == models.Booking.room_id).all()

    # 3. 构造数据并同步更新数据库
    data = []
    for b in bookings:
        display_status = b.status

        # --- 核心逻辑：自动判定并同步数据库 ---
        if b.status == "待入住" and b.end_date < today:
            display_status = "已离店/完成"
            # 💡 这里直接更新数据库，确保前端页面也同步变掉
            db.query(models.Booking).filter(models.Booking.id == b.id).update({"status": "已离店/完成"})
        elif b.status == "待入住" and b.start_date <= today <= b.end_date:
            display_status = "入住中"
            db.query(models.Booking).filter(models.Booking.id == b.id).update({"status": "入住中"})

        data.append({
            "订单编号": b.id,
            "顾客姓名": b.guest_name,
            "房间号": b.room_number,
            "入住日期": b.start_date.strftime("%Y-%m-%d") if b.start_date else "",
            "离店日期": b.end_date.strftime("%Y-%m-%d") if b.end_date else "",
            "订单状态": display_status
        })

    # 提交数据库的所有更改
    db.commit()

    df = pd.DataFrame(data)

    # 4. 写入 Excel 并应用美化样式
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='营收报表')
        worksheet = writer.sheets['营收报表']

        # 蓝色表头样式
        header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        for i in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(i)
            # 设置宽度为 25 像素，确保日期清晰
            worksheet.column_dimensions[col_letter].width = 25

            cell = worksheet.cell(row=1, column=i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

            # 给数据行也加上居中对齐，显得更整齐
            for row_idx in range(2, len(data) + 2):
                worksheet.cell(row=row_idx, column=i).alignment = center_align

    output.seek(0)

    # 5. 生成动态文件名下载
    timestamp = datetime.now().strftime('%H%M%S')
    filename = f"酒店营收分析_{timestamp}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}",
            "Cache-Control": "no-cache"
        }
    )


# 1. 获取所有会员
@app.get("/api/members", response_model=List[schemas.Member])
def get_members(db: Session = Depends(get_db)):
    return db.query(models.Member).all()


# 2. 注册/新增会员
@app.post("/api/members", response_model=schemas.Member)
def create_member(member: schemas.MemberCreate, db: Session = Depends(get_db)):
    db_member = models.Member(**member.dict())
    db.add(db_member)
    db.commit()
    db.refresh(db_member)
    return db_member


# 3. 更新会员信息 (改等级、充值、加积分)
@app.put("/api/members/{member_id}")
def update_member(member_id: int, data: schemas.MemberUpdate, db: Session = Depends(get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="会员不存在")

    # 动态更新字段
    update_data = data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_member, key, value)

    db.commit()
    db.refresh(db_member)
    return {"message": "更新成功", "data": db_member}


# 4. 删除会员
@app.delete("/api/members/{member_id}")
def delete_member(member_id: int, db: Session = Depends(get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="会员不存在")
    db.delete(db_member)
    db.commit()
    return {"message": "会员已注销"}


# 5. 会员结算与结账接口
@app.post("/api/bookings/{booking_id}/checkout")
def checkout_booking(booking_id: int, db: Session = Depends(get_db)):
    # 1. 查找订单
    booking = db.query(models.Booking).filter(models.Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="订单未找到")

    # 2. 查找关联房间
    room = db.query(models.Room).filter(models.Room.id == booking.room_id).first()

    # 3. 计算折扣逻辑 (假设根据手机号匹配会员)
    # 这里我们尝试在数据库寻找同名的会员，或者你可以让前端传个 member_id 过来
    member = db.query(models.Member).filter(models.Member.name == booking.guest_name).first()

    original_price = room.price
    final_price = original_price
    discount_msg = "原价结算"

    if member:
        # 根据等级打折
        if member.level == "钻石会员":
            final_price = original_price * 0.8  # 8折
            discount_msg = "钻石会员 8 折优惠"
        elif member.level == "白金会员":
            final_price = original_price * 0.9  # 9折
            discount_msg = "白金会员 9 折优惠"

        # 增加积分 (1元=1分)
        member.points += int(final_price)
        # 记录积分流水
        db.add(models.MemberLog(
            member_id=member.id,
            type="积分",
            amount=float(int(final_price)),
            reason=f"房间 {room.number} 结账获得积分"
        ))

    # 4. 更新订单和房间状态
    booking.status = "已离店/完成"
    booking.actual_revenue = final_price  # 记录实际营收
    room.status = "空闲"
    room.guest_name = None  # 清空房间住客信息

    db.commit()

    return {
        "message": "结账成功",
        "original_price": original_price,
        "final_price": final_price,
        "discount_info": discount_msg,
        "points_earned": int(final_price) if member else 0
    }


# 6. 提交房间评价
@app.post("/api/comments")
def create_comment(comment: schemas.CommentBase, db: Session = Depends(get_db)):
    new_comment = models.Comment(**comment.dict())
    db.add(new_comment)
    db.commit()
    db.refresh(new_comment)
    return {"message": "感谢您的评价！", "data": new_comment}

# 7. 获取某个房间的所有评价
@app.get("/api/rooms/{room_id}/comments")
def get_room_comments(room_id: int, db: Session = Depends(get_db)):
    return db.query(models.Comment).filter(models.Comment.room_id == room_id).all()


# 处理实名入住，并将信息写入 Room 表
@app.post("/api/bookings/{booking_id}/checkin")
def checkin_booking(booking_id: int, data: schemas.RoomStatusUpdate, db: Session = Depends(get_db)):
    # 1. 查找订单
    booking = db.query(models.Booking).filter(models.Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="未找到预订订单")

    # 2. 查找关联房间
    room = db.query(models.Room).filter(models.Room.id == booking.room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="未找到关联房间")

    # 3. 更新订单状态
    booking.status = "入住中"

    # 4. 【核心修复】将实名信息同步到房间表
    # 这样 Dashboard 页面请求 /api/rooms 时才能拿到这些值
    room.status = "已入住"
    room.guest_name = data.guest_name
    room.guest_id_card = data.guest_id_card
    room.guest_phone = data.guest_phone

    db.commit()
    return {"message": "实名入住办理成功"}


@app.post("/api/rooms/{room_id}/walk-in")
def room_walk_in(room_id: int, request: schemas.WalkInRequest, db: Session = Depends(get_db)):
    # 注意上面的类型注解变成了 schemas.WalkInRequest

    # 1. 获取房间
    room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")
    if room.status == "已入住":
        raise HTTPException(status_code=400, detail="该房间已有客人")

    # 2. 创建订单记录
    new_booking = models.Booking(
        room_id=room.id,
        guest_name=request.guest_name,
        start_date=request.check_in_date,
        end_date=request.check_out_date,
        status="入住中"
    )
    db.add(new_booking)

    # 3. 更新房间物理状态
    room.status = "已入住"
    room.guest_name = request.guest_name
    room.guest_id_card = request.guest_id_card
    room.guest_phone = request.guest_phone

    db.commit()
    return {"message": "入住办理成功", "booking_id": new_booking.id}


# 2. 预结账信息 (获取退房账单)
@app.get("/api/rooms/{room_id}/bill")
def get_room_bill(room_id: int, db: Session = Depends(get_db)):
    room = db.query(models.Room).filter(models.Room.id == room_id).first()
    if room.status != "已入住":
        raise HTTPException(status_code=400, detail="该房间未入住，无法结账")

    # 找到该房间当前正在进行的订单
    # 逻辑：查找关联该房间，且状态不是“已离店”的最新订单
    booking = db.query(models.Booking).filter(
        models.Booking.room_id == room_id,
        models.Booking.status == "入住中"
    ).order_by(models.Booking.id.desc()).first()

    if not booking:
        # 如果找不到订单（可能是老数据），就按1天计算
        days = 1
        booking_id = 0
    else:
        # 计算实际入住天数
        today = date.today()
        # 如果入住日期是今天，按1天算；否则按实际差值算
        delta = (today - booking.start_date).days
        days = delta if delta > 0 else 1
        booking_id = booking.id

    total_amount = days * room.price

    return {
        "room_number": room.number,
        "room_type": room.room_type,
        "guest_name": room.guest_name,
        "price_per_night": room.price,
        "stay_days": days,
        "total_amount": total_amount,
        "booking_id": booking_id  # 返回订单ID方便后续结账
    }


# 3. 确认结账退房
@app.post("/api/bookings/{booking_id}/confirm-checkout")
def confirm_checkout(booking_id: int, db: Session = Depends(get_db)):
    booking = db.query(models.Booking).filter(models.Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="订单不存在")

    # 更新订单
    booking.status = "已离店/完成"
    # 这里可以加入 actual_revenue 更新逻辑

    # 释放房间
    room = db.query(models.Room).filter(models.Room.id == booking.room_id).first()
    room.status = "待打扫"  # 结账后变为待打扫
    room.guest_name = None
    room.guest_id_card = None
    room.guest_phone = None

    db.commit()
    return {"message": "退房结账成功"}
